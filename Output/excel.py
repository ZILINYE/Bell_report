import pandas
from openpyxl import load_workbook


def Main(data, sheetname):
    book = load_workbook('Output/report.xlsx')
    writer = pandas.ExcelWriter('Output/report.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, sheetname, startrow=2, header=False)
    writer.save()


def Other(data, time_list, sheetname):
    header_list = ['C1', 'E1', 'G1', 'I1', 'K1', 'M1']
    dest_filename = "Output/report.xlsx"
    book = load_workbook(dest_filename)
    writer = pandas.ExcelWriter(dest_filename, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, sheetname, startrow=1, header=True)
    writer.save()
    ws1 = book[sheetname]
    time_list = time_list['week_day']
    time_length = len(time_list) / 4
    i = 1
    while i <= time_length:
        start_hh = time_list[4 * i - 4]
        start_mm = time_list[4 * i - 3]
        end_hh = time_list[4 * i - 2]
        end_mm = time_list[4 * i - 1]
        time_range = str(start_hh) + ':' + str(start_mm) + '~' + str(end_hh) + ':' + str(end_mm)
        ws1[header_list[i - 1]] = time_range
        i += 1
    writer.save()
